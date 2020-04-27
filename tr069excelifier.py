"""
Copyright 2020 Mikko Jaakkola (mikko.la.jaakkola@gmail.com)
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"),
to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense,
and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

import argparse
import re
import xml.etree.ElementTree as ET

from xml.etree.ElementTree import Element
from typing import Dict, Tuple, List

import pandas as pd

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def get_params(obj: Element) -> Tuple[str,str,str]:
    """Parses parameters from management object

    Arguments:
        obj {Element} -- Parameter Element

    Returns:
        str -- All the parameters field as single string
    """

    def add_optionals(prev: str, param: Element) -> Dict[str,str]:
        """Helper function for adding optional elements

        Arguments:
            prev {str} -- The parameter string so far and where the new values will be concatanated
            param {Element} -- Element object for parameter element

        Returns:
            Dict[str,str] -- Object parameter attributes
        """
        optional = obj.get(param)
        if optional:
            prev = "{} {:<8} ".format(prev, optional.strip(' \n'))
        return prev

    desc = re.sub(' +', ' ', obj.find('description').text.strip('\n'))

    text = ""
    text = add_optionals(text, 'status')
    text = add_optionals(text, 'activeNotify')
    text = add_optionals(text, 'forcedInform')

    return (
        obj.get('name').strip('\n'),
        obj.get('access').strip('\n'),
        f"{text} {desc}"
    )

def parse_object(obj: Element, version: str) -> List[Dict[str, str]]:
    """Parses device model object

    Arguments:
        obj {Element} -- Device object
        version {str} -- Device model version

    Returns:
        List[Dict[str, str]] -- List of object dictionaries from the same root.
    """
    combined = []

    params = obj.findall('parameter')
    if params:
        for param in params:
            (name, access, desc) = get_params(param)

            combined.append({
                'Object' : obj.get('name') + name,
                'Access': obj.get('access'),
                'Description': re.sub(' +', ' ', obj.find('description').text.strip('\n')),
                'Model': version,
                'Parameter Access': access,
                'Parameter Description' : desc
            })
    else:
        # We can end up here in hierarchical objects
        combined.append({
            'Object' : obj.get('name'),
            'Access': obj.get('access'),
            'Description': re.sub(' +', ' ', obj.find('description').text.strip('\n')),
            'Model': version,
            'Parameter Access': "",
            'Parameter Description' : ""
        })

    return combined


# Parsing profiles
def get_profile_params(obj: Element) -> str:
    """Acquires parameters from profile reference

    Arguments:
        obj {Element} -- Profile object reference

    Returns:
        str -- Concatanated string of all parameters
    """
    name = obj.get('ref').strip('\n')
    access = obj.get('requirement').strip('\n')

    text = "{:<12} ".format(access)

    return text + name

def parse_profile(obj: Element, profile: Element) -> Dict[str,str]:
    """Parses model profile and related object element.

    Function is called once per object element and profile information is repeated per object.

    Arguments:
        obj {Element} -- Profile object
        profile {Element} -- Profile

    Returns:
        Dict[str,str] -- Profile object dictionary
    """

    obj_dict = {
        'Profile': profile.get('name'),
        'Name' : obj.get('ref'),
        'Requirement': obj.get('requirement'),
        'Base': profile.get('base') or "",
        'Extends': profile.get('extends') or ""
    }

    params = obj.findall('parameter')
    if params:
        obj_dict['Parameters'] = "\n".join([get_profile_params(param) for param in params])

    return obj_dict


def build_sheet(ws: Worksheet, data: pd.DataFrame, columns: List[str]):
    """Common worksheet builder function to harmonize style settings

    Arguments:
        ws {Worksheet} -- Worksheet to be formatted
        data {pd.DataFrame} -- Data to be inserted into Worksheet
        columns {List[str]} -- Columns that shall be used as bulk-text with wrapping and etc.
    """
    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

    font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = font

    alignment = Alignment(horizontal='general',vertical='top',wrap_text=True)

    for column in columns:
        for cell in ws[f"{column}:{column}"]:
            cell.alignment = alignment


def parse_model(filename: str, output: str):
    """Parses Device model XML file and exports the data into Excel

    Arguments:
        filename {str} -- Input XML filename
        output {str} -- Output Excel filename
    """
    tree = ET.parse(filename)
    #I'm only interested in the model and profiles so the rest is discarded for now
    model = tree.getroot().find('model')

    # These build both model and profile parts
    objects = [final_obj for obj in model.findall('object') for final_obj in parse_object(obj, model.get('name'))]
    profiles = [parse_profile(obj, profile) for profile in model.findall('profile') for obj in profile.findall('object')]

    df_model = pd.DataFrame.from_records(objects)
    df_profile = pd.DataFrame.from_records(profiles).sort_values(by=['Profile', 'Name', 'Base'])

    # We have both model and profiles in the dataframes so from here we can just generate Excel
    wb = Workbook()
    ws_model = wb.active
    ws_model.title = "Model"
    ws_profile = wb.create_sheet("Profiles")

    build_sheet(ws_model, df_model, ['D', 'E'])
    build_sheet(ws_profile, df_profile, ['F'])

    # Formatting the column widths roughly right. Not pretty but will do for now.
    ws_model.column_dimensions['A'].width=75.0
    ws_model.column_dimensions['B'].width=12.0
    ws_model.column_dimensions['C'].width=40.0
    ws_model.column_dimensions['D'].width=12.0
    ws_model.column_dimensions['E'].width=15.0
    ws_model.column_dimensions['F'].width=400.0

    ws_profile.column_dimensions['A'].width=25.0
    ws_profile.column_dimensions['B'].width=50.0
    ws_profile.column_dimensions['C'].width=10.0
    ws_profile.column_dimensions['D'].width=20.0
    ws_profile.column_dimensions['E'].width=20.0
    ws_profile.column_dimensions['F'].width=400.0
    wb.save(output)

if __name__=='__main__':
    parser = argparse.ArgumentParser(description='Converts TR069 XML models into Excel')
    parser.add_argument('-f', '--file', help="Input XML model file", required=True)
    parser.add_argument('-o', '--output', help="Output file (should have .xlsx postfix)", default="output.xlsx")
    args = parser.parse_args()
    parse_model(args.file, args.output)
