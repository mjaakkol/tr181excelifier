"""
Copyright 2020 Mikko Jaakkola (mikko.la.jaakkola@gmail.com)
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"),
to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense,
and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

import argparse
import re
import xml.etree.ElementTree as ET

from xml.etree.ElementTree import Element
from typing import Dict, Tuple, List

import pandas as pd

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def get_params(obj: Element) -> Tuple[str,str,str]:
    """Parses parameters from management object

    Arguments:
        obj {Element} -- Parameter Element

    Returns:
        str -- All the parameters field as single string
    """

    def add_optionals(prev: str, param_name: str) -> str:
        """Helper function for adding optional elements

        Arguments:
            prev {str} -- The parameter string so far and where the new values will be concatanated
            param {str} -- Element object for parameter element

        Returns:
            Tuple
            str -- Object parameter name
            str -- Object parameter access attribute
            str -- Object parameter description
        """
        optional = obj.get(param_name)
        if optional:
            if param_name != 'syntax':
                prev = "{} {:<8} ".format(prev, optional)
        return prev

    def add_syntax(prev: str, desc: str) -> Tuple[str,str]:
        """Adds information from syntax tag containing parameter type information

        Arguments:
            prev {str} -- Input string to be used as bases to append new items
            desc {str} -- Description string where applicable tags are filled from found items

        Returns:
            str -- Constructed string
            desc -- Description string with potential replacements
        """
        syntax = obj.find('syntax')
        if syntax:
            syntax_text = ""
            param_type = syntax.find('boolean')

            if not param_type:
                param_type = syntax.find('string')

                if not param_type:
                    # We'll take the first item (this is kind of ugly as relies into ordering)
                    param_type = list(syntax)
                    if param_type:
                        param_type = param_type[0]
                        syntax_text = param_type.tag
                        units_tag = param_type.find('units')

                        if units_tag is not None:
                            unit_value = units_tag.get('value')
                            syntax_text = "{} in {}".format(syntax_text, unit_value)
                            desc = re.sub('{{units}}', unit_value, desc)

                        prev = "{} {}".format(prev, syntax_text)
                    else:
                        # We can hit here in some non-standard types and that's OK
                        pass

                else:
                    # String parameter processing
                    enum_type = param_type.findall('enumeration')
                    if enum_type:
                        prev = "Enums ({})".format("|".join([e.get('value') for e in enum_type]))
                    else:
                        param_type = param_type.find('size')

                        if param_type:
                            syntax_text = "max length " + param_type.get('maxLength')

                        prev = f"{prev} String{syntax_text}"

            else:
                prev = f"{prev} Boolean"

            param_type = syntax.find('default')
            if param_type:
                prev = "{} {}".format(prev,param_type.text)

        return prev, desc

    desc = obj.find('description').text

    text = ""
    text = add_optionals(text, 'status')
    text = add_optionals(text, 'activeNotify')
    text = add_optionals(text, 'forcedInform')
    (text, desc) = add_syntax(text, desc)

    return (
        obj.get('name'),
        obj.get('access'),
        f"{text} {desc}"
    )

def parse_object(obj: Element) -> List[Dict[str, str]]:
    """Parses device model object

    Arguments:
        obj {Element} -- Device object

    Returns:
        List[Dict[str, str]] -- List of object dictionaries from the same root.
    """
    combined = []

    params = obj.findall('parameter')
    if params:
        for param in params:
            (name, access, desc) = get_params(param)

            combined.append({
                'Object' : obj.get('name'),
                'Access': obj.get('access'),
                'Description': re.sub(' +', ' ', obj.find('description').text),
                'Parameter': name,
                'Parameter Access': access,
                'Parameter Description' : desc
            })
    else:
        # We can end up here in hierarchical objects
        combined.append({
            'Object' : obj.get('name'),
            'Access': obj.get('access'),
            'Description': re.sub(' +', ' ', obj.find('description').text),
            'Parameter': "",
            'Parameter Access': "",
            'Parameter Description' : ""
        })

    return combined


# Parsing profiles
def get_profile_params(obj: Element) -> str:
    """Acquires parameters from profile reference

    Arguments:
        obj {Element} -- Profile object reference

    Returns:
        str -- Concatanated string of all parameters
    """
    name = obj.get('ref')
    access = obj.get('requirement')
    return "{:<12} {}".format(access, name)

def parse_profile(obj: Element, profile: Element) -> Dict[str,str]:
    """Parses model profile and related object element.

    Function is called once per object element and profile information is repeated per object.

    Arguments:
        obj {Element} -- Profile object
        profile {Element} -- Profile

    Returns:
        Dict[str,str] -- Profile object dictionary
    """

    obj_dict = {
        'Profile': profile.get('name'),
        'Name' : obj.get('ref'),
        'Requirement': obj.get('requirement'),
        'Base': profile.get('base') or "",
        'Extends': profile.get('extends') or ""
    }

    params = obj.findall('parameter')
    if params:
        obj_dict['Parameters'] = "\n".join([get_profile_params(param) for param in params])

    return obj_dict


def build_sheet(ws: Worksheet, data: pd.DataFrame, columns: List[str]):
    """Common worksheet builder function to harmonize style settings

    Arguments:
        ws {Worksheet} -- Worksheet to be formatted
        data {pd.DataFrame} -- Data to be inserted into Worksheet
        columns {List[str]} -- Columns that shall be used as bulk-text with wrapping and etc.
    """
    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

    font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = font

    alignment = Alignment(horizontal='general',vertical='top',wrap_text=True)

    for column in columns:
        for cell in ws[f"{column}:{column}"]:
            cell.alignment = alignment

def clean_model(model: pd.DataFrame) -> pd.DataFrame:
    """Cleans all markdown language and unnecessary characters from the fields

    Arguments:
        model {pd.DataFrame} --Input DataFrame for cleaning

    Returns:
        pd.DataFrame -- Cleaned DataFrame
    """
    def process_text(text) -> str:
        """Process all the elements for the common operations like stripping extra spaces and newlines

        Arguments:
            text {str} -- Text of the cell to be manipulated

        Returns:
            str -- The final string without extra spaces and newlines
        """
        text = re.sub(' +', ' ', text)
        return re.sub('\n', '', text).strip()

    def map_params(row: pd.Series, column: str) -> str:
        """Cleans description strings from markdown language a row at the time.

        Arguments:
            row {pd.Series} -- Row where to apply the transformations
            column {str} -- Name of the column to be manipulated

        Returns:
            str -- The cleaned description string
        """
        # Replaces the references to the current object and param with the actual name.
        text = re.sub('{{object}}', row['Object'], row[column])
        text = re.sub('{{param}}', row['Parameter'], text)

        # Basic stuff not really helping us
        filtered = r'( {{numentries}}| {{datatype\|expand}}| {{pattern}}| {{enum}}| {{list}}| {{reference}}| {{noreference}})'
        text = re.sub(filtered, r'', text)

        # Removing text|preserved strings
        text = re.sub(r'{{\w+\|([^}]+)}}', r'\1', text)

        # Removing common {{word}} items. This must be the last item as it can remove something usefull as well
        text = re.sub(r'{{(\w+)}}', r'\1', text)
        return text

    # This is pretty heavy and done for each item. We maybe able to remove this as most the undesired characters are
    # in the description parts
    model = model.applymap(process_text)

    model['Parameter Description'] = model.apply(lambda x: map_params(x,'Parameter Description'), axis=1)
    model['Description'] = model.apply(lambda x: map_params(x,'Description'), axis=1)

    return model


def parse_model(filename: str, output: str):
    """Parses Device model XML file and exports the data into Excel

    Arguments:
        filename {str} -- Input XML filename
        output {str} -- Output Excel filename
    """
    tree = ET.parse(filename)
    #I'm only interested in the model and profiles so the rest is discarded for now
    model = tree.getroot().find('model')
    version = model.get('name')

    # These build both model and profile parts
    objects = [final_obj for obj in model.findall('object') for final_obj in parse_object(obj)]
    profiles = [parse_profile(obj, profile) for profile in model.findall('profile') for obj in profile.findall('object')]

    df_model = clean_model(pd.DataFrame.from_records(objects))

    unique_items = pd.unique(df_model['Object'])
    index_ranges_list = [(2+min(df_model[df_model['Object'] == o].index), 2+max(df_model[df_model['Object'] == o].index)) for o in unique_items]

    df_model.rename(index={0: f"Object ({version})"}, inplace=True)

    df_profile = pd.DataFrame.from_records(profiles).sort_values(by=['Profile', 'Name', 'Base'])

    # We have both model and profiles in the dataframes so from here we can just generate Excel
    wb = Workbook()
    ws_model = wb.active
    ws_model.title = "Model"
    ws_profile = wb.create_sheet("Profiles")

    build_sheet(ws_model, df_model, ['C', 'D', 'E'])
    build_sheet(ws_profile, df_profile, ['F'])

    # Performing cell merges where applicable
    for (min_cell, max_cell) in index_ranges_list:
        for column in 'ABC':
            ws_model.merge_cells(f'{column}{min_cell}:{column}{max_cell}')

    # Formatting the column widths roughly right. Not pretty but will do for now.
    ws_model.column_dimensions['A'].width=50.0
    ws_model.column_dimensions['B'].width=11.0
    ws_model.column_dimensions['C'].width=40.0
    ws_model.column_dimensions['D'].width=45.0
    ws_model.column_dimensions['E'].width=15.0
    ws_model.column_dimensions['F'].width=400.0

    ws_profile.column_dimensions['A'].width=25.0
    ws_profile.column_dimensions['B'].width=50.0
    ws_profile.column_dimensions['C'].width=10.0
    ws_profile.column_dimensions['D'].width=20.0
    ws_profile.column_dimensions['E'].width=20.0
    ws_profile.column_dimensions['F'].width=400.0
    wb.save(output)

if __name__=='__main__':
    parser = argparse.ArgumentParser(description='Converts TR069 XML models into Excel')
    parser.add_argument('-f', '--file', help="Input XML model file", required=True)
    parser.add_argument('-o', '--output', help="Output file (should have .xlsx postfix)", default="output.xlsx")
    args = parser.parse_args()
    parse_model(args.file, args.output)
